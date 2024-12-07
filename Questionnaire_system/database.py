from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import random
import string

db = SQLAlchemy()

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # 'teacher' 或 'student'
    student_id = db.Column(db.String(7), unique=True, nullable=True)  # 学生学号，7位
    real_name = db.Column(db.String(50), nullable=True)  # 真实姓名
    
    # 添加关系
    surveys = db.relationship('Survey', backref='teacher', lazy=True)
    responses = db.relationship('Response', backref='student', lazy=True)
    
class Survey(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)  # 控制问卷是否开放
    code = db.Column(db.String(10), unique=True, nullable=True)  # 用于二维码的唯一代码
    
    # 添加关系
    questions = db.relationship('Question', backref='survey', lazy=True, cascade='all, delete-orphan')
    responses = db.relationship('Response', backref='survey', lazy=True)
    
    def get_statistics(self):
        """获取问卷的统计数据"""
        stats = {}
        # 获取所有提交了该问卷的学生数量
        total_responses = SurveyResponse.query.filter_by(survey_id=self.id).count()
        
        for question in self.questions:
            question_stats = {
                'question_text': question.question_text,
                'question_type': question.question_type,
                'total_responses': total_responses,
                'options': []
            }
            
            if question.question_type in ['single', 'multiple']:
                # 获取所有有效的问卷响应ID
                survey_responses = SurveyResponse.query.filter_by(survey_id=self.id).all()
                survey_response_ids = [sr.id for sr in survey_responses]
                
                # 获取所有回答了这个问题的响应
                responses = Response.query.filter(
                    Response.survey_id == self.id,
                    Response.question_id == question.id,
                    Response.survey_response_id.in_(survey_response_ids)
                ).all()
                
                if question.question_type == 'single':
                    # 单选题：基于回答人数计算百分比
                    responding_students = len(set(r.survey_response_id for r in responses))
                    
                    # 统计每个选项的选择情况
                    for option in question.options:
                        # 统计选择了该选项的次数
                        option_selections = Response.query.filter(
                            Response.survey_id == self.id,
                            Response.question_id == question.id,
                            Response.option_id == option.id,
                            Response.survey_response_id.in_(survey_response_ids)
                        ).count()
                        
                        # 计算百分比（基于回答了问题的学生数量）
                        percentage = (option_selections / responding_students * 100) if responding_students > 0 else 0
                        
                        question_stats['options'].append({
                            'option_text': option.option_text,
                            'count': option_selections,
                            'percentage': round(percentage, 2)
                        })
                else:
                    # 多选题：基于总选择次数计算百分比
                    total_selections = len(responses)  # 所有选项被选择的总次数
                    
                    # 统计每个选项的选择情况
                    for option in question.options:
                        # 统计选择了该选项的次数
                        option_selections = Response.query.filter(
                            Response.survey_id == self.id,
                            Response.question_id == question.id,
                            Response.option_id == option.id,
                            Response.survey_response_id.in_(survey_response_ids)
                        ).count()
                        
                        # 计算百分比（基于总选择次数）
                        percentage = (option_selections / total_selections * 100) if total_selections > 0 else 0
                        
                        question_stats['options'].append({
                            'option_text': option.option_text,
                            'count': option_selections,
                            'percentage': round(percentage, 2)
                        })
                
            elif question.question_type == 'text':
                # 对于文本题的处理
                text_responses = Response.query.filter(
                    Response.survey_id == self.id,
                    Response.question_id == question.id,
                    Response.survey_response_id.in_([sr.id for sr in SurveyResponse.query.filter_by(survey_id=self.id)])
                ).all()
                
                question_stats['text_answers'] = [
                    {
                        'answer': response.text_answer,
                        'student': response.student.username,
                        'student_id': response.student.student_id
                    } for response in text_responses
                ]
            
            stats[question.id] = question_stats
            
        return stats
    
    def to_excel(self):
        """导出问卷数据为Excel格式"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "问卷回答统计"
        
        # 设置标题样式
        title_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 写入问卷标题
        ws['A1'] = self.title
        ws['A1'].font = title_font
        
        # 写入表头
        headers = ['提交时间', '学生姓名']
        current_col = 3
        question_columns = {}  # 记录每个问题的列号
        
        for question in self.questions:
            col_letter = ws.cell(row=2, column=current_col).column_letter
            question_columns[question.id] = col_letter
            ws[f'{col_letter}2'] = question.question_text
            ws[f'{col_letter}2'].fill = header_fill
            current_col += 1
        
        # 写入数据
        current_row = 3
        for sr in self.survey_responses:
            ws.cell(row=current_row, column=1, value=sr.submitted_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=current_row, column=2, value=sr.student.username)
            
            # 获取该学生的所有回答
            responses = Response.query.filter_by(
                student_id=sr.student_id,
                survey_id=self.id
            ).all()
            
            for response in responses:
                col_letter = question_columns[response.question_id]
                ws[f'{col_letter}{current_row}'] = response.selected_option.option_text
                
            current_row += 1
            
        return wb
    
    def generate_code(self):
        """生成唯一的问卷代码"""
        while True:
            # 生成8位随机字符串
            code = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
            # 检查是否已存在
            if not Survey.query.filter_by(code=code).first():
                self.code = code
                break

class Question(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    survey_id = db.Column(db.Integer, db.ForeignKey('survey.id'), nullable=False)
    question_text = db.Column(db.String(500), nullable=False)
    question_type = db.Column(db.String(20), default='single')  # 'single', 'multiple' 或 'text'
    order = db.Column(db.Integer, nullable=False)
    
    options = db.relationship('Option', backref='question', lazy=True, cascade='all, delete-orphan')
    responses = db.relationship('Response', backref='question', lazy=True)

class Option(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    question_id = db.Column(db.Integer, db.ForeignKey('question.id'), nullable=False)
    option_text = db.Column(db.String(200), nullable=False)
    order = db.Column(db.Integer, nullable=False)  # 选项顺序
    
    # 添加关系
    responses = db.relationship('Response', backref='selected_option', lazy=True)

class Response(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    survey_id = db.Column(db.Integer, db.ForeignKey('survey.id'), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey('question.id'), nullable=False)
    option_id = db.Column(db.Integer, db.ForeignKey('option.id'), nullable=True)  # 允许为空,用于文本答案
    text_answer = db.Column(db.Text, nullable=True)  # 新增:文本答案
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)
    survey_response_id = db.Column(db.Integer, db.ForeignKey('survey_response.id'), nullable=False)  # 新增字段

class SurveyResponse(db.Model):
    """用于记录整份问卷的提交记录"""
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    survey_id = db.Column(db.Integer, db.ForeignKey('survey.id'), nullable=False)
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # 添加关系
    student = db.relationship('User', backref='survey_responses')
    survey = db.relationship('Survey', backref='survey_responses')
    
class Feedback(db.Model):
    """用于记录学生意见反馈"""
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='unread')  # 状态：unread（未）, read（已读）
    reply = db.Column(db.Text, nullable=True)  # 教师回复内容
    reply_at = db.Column(db.DateTime, nullable=True)  # 回复时间
    
    # 添加关系
    student = db.relationship('User', backref='feedbacks')
    
class FeedbackReply(db.Model):
    """用于记录教师对反馈的回复"""
    id = db.Column(db.Integer, primary_key=True)
    feedback_id = db.Column(db.Integer, db.ForeignKey('feedback.id'), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # 添加关系
    feedback = db.relationship('Feedback', backref='replies')
    